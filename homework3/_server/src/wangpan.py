import os,re,time,json,socket
import datetime

from openpyxl import load_workbook
import logging
logging.basicConfig(level=logging.DEBUG,
                    format='%(asctime)s - %(filename)s[line:%(lineno)d] - %(levelname)s: %(message)s')  # logging.basicConfig函数对日志的输出格式及方式做相关配置
# 由于日志基本配置中级别设置为DEBUG，所以一下打印信息将会全部显示在控制台上
from _server.utils import req
from _server._config import common

class Pan(object):
    def __init__(self,conn):
        self.conn = conn
        self.username = None

    @property
    def home_path(self): # 为用户设置独立文件夹
        return os.path.join(common.USER_FOLDER_PATH,self.username)

    def send_json_data(self, **kwargs):
        req.send_data(self.conn, json.dumps(kwargs))

    def recv_save_file(self, target_file_path):
        req.recv_save_file(self.conn, target_file_path)

    def login(self,username,password):
        wb = load_workbook(common.DB_FILE_PATH)
        sheet = wb.worksheets[0]
        success = False
        for row in sheet.iter_rows(2):
            if username == row[0].value and password == row[1].value :
                success = True
                break
        if success:
            logging.info(self.send_json_data(status=True, data="登录成功"))
            self.username = username
        else:
            logging.error(
                self.send_json_data(status=False, error="登录失败"))


    def register(self, username, password):
        """ 用户注册， 用户名和密码写入到excel中（已存在则不再注册） """
        wb = load_workbook(common.DB_FILE_PATH)
        sheet = wb.worksheets[0]
        exist = False
        for row in sheet.rows:
            if row[0] == username:
                exist = True
        if exist:
            logging.error(self.send_json_data(status=False, error="用户名已存在"))
            return
        _maxrow = sheet.max_row
        data_list = [username, password, datetime.datetime.now().strftime("%Y-%m-%d")]
        for index,item in enumerate(data_list,1):
            cell = sheet.cell(_maxrow+1,index)
            cell.value = item
        wb.save(common.DB_FILE_PATH)
        user_dir = os.path.join(common.USER_FOLDER_PATH,username)
        os.mkdir(user_dir)
        logging.info(
            self.send_json_data(status=True, data="注册成功"))

    def ls(self,fold_path = None):
        if not self.username:
            logging.warning(self.send_json_data(status=False, error="登录后才能查看"))
            return
        target_folder = os.path.join(self.home_path,fold_path)

        if not os.path.exists(target_folder):
            logging.error(self.send_json_data(status=False, error="路径不存在"))
            return
        if not os.path.isdir(target_folder):
            logging.info(
                self.send_json_data(status=False, error="文件夹不存在"))
            return

        data = "\n".join(os.listdir(target_folder))
        self.send_json_data(status=True, data=data)
    def upload(self,file_path):
        if not self.username:
            logging.warning(
                self.send_json_data(status=False, error="登录后才能查看"))
            return

        target_file_path = os.path.join(self.home_path, file_path)
        folder = os.path.dirname(target_file_path)
        if not os.path.exists(folder):
            os.makedirs(folder)

        logging.info(
            self.send_json_data(status=True, data="开始上传"))

        # *** 很关键，如果是基于 select + 非阻塞的模式下，客户度未发送数据，直接去recv，会出现blocking error的错误 ***
        # 解决方案：1.改为阻塞模式； 2.使用time.sleep等一会，等有数据了再去读取。
        # time.sleep(2)

        # 接收上传的文件
        self.recv_save_file(target_file_path)

    def download(self,file_path, seek =0 ):
        """ 下载文件，支持断点续传（客户端本地已有文件）
                    seek=None，从头开始下载；
                    seek=1000，从1000字节开始下载（续传）
                """
        # 用户未登录
        if not self.username:
            logging.warning(
                self.send_json_data(status=False, error="登录成功后才能上传"))
            return

        target_file_path = os.path.join(self.home_path, file_path)
        if not os.path.exists(target_file_path):
            # req.send_data(self.conn, json.dumps({"status": False, "error": "文件{}不存在".format(file_path)}))
            logging.error(self.send_json_data(status=False, error="文件{}不存在".format(file_path)))
            return
            # 获取文件大小并返回
            # req.send_data(self.conn, json.dumps({"status": True, "data": "开始下载"}))
        logging.info(self.send_json_data(status=True, data="开始下载"))

        # 发送文件
        seek = int(seek)
        total_size = os.stat(target_file_path).st_size
        req.send_file_BySeek(self.conn, total_size - seek, target_file_path, seek)


    def _execute(self):
        conn = self.conn
        cmd = req.recv_data(conn).decode('utf8')
        if cmd.upper() == "Q":
            logging.info(("客户端退出"))
            return False
        method_map = {
            "login": self.login,
            "register": self.register,
            "ls": self.ls,
            "upload": self.upload,
            "download": self.download,
        }
        cmd, *args = re.split(r"\s+", cmd)  # [register,root,123] /  [login,root,123]
        method = method_map[cmd]
        method(*args)

        return True


